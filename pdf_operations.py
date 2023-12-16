#pdf_operations.py
import pandas as pd
import os
import glob
import pdf2image
import pdf2docx
import PyPDF2
import tempfile
import tabula
import openpyxl
import warnings
import PIL

def pdf_to_excel(input_pdf_path):
    # Define output Excel file path
    output_excel_path = os.path.splitext(input_pdf_path)[0] + ".xlsx"

    # Create a temporary directory for storing temporary CSV files
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_csv_files = []

        # Convert each page of the PDF to a separate CSV file
        pdf_reader = PyPDF2.PdfReader(input_pdf_path)
        num_pages = len(pdf_reader.pages)

        for page_num in range(num_pages):
            temp_csv_file = os.path.join(temp_dir, f"Page{page_num + 1}.csv")
            temp_csv_files.append(temp_csv_file)

            try:
                tabula.convert_into(input_pdf_path, temp_csv_file, output_format="csv", pages=page_num + 1)
            except Exception as e:
                print(f"An error occurred while converting page {page_num + 1} to CSV: {e}")

        # Merge all the separate CSV files into one DataFrame
        dfs = []
        for temp_csv_file in temp_csv_files:
            df = pd.read_csv(temp_csv_file)
            dfs.append(df)

        # Concatenate all DataFrames into a single DataFrame, skipping the title of every page after the first page
        merged_data_frame = dfs[0]  # Initialize with the first page
        for df in dfs[1:]:  # Skip the first page's title
            merged_data_frame = pd.concat([merged_data_frame, df], ignore_index=True)

        # Save the merged DataFrame to Excel
        merged_data_frame.to_excel(output_excel_path, index=False)

        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(output_excel_path)
        sheet = workbook.active

        # Iterate through columns
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)

            # Find the maximum length in each column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            # Adjust the column width
            adjusted_width = (max_length + 3)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(output_excel_path)
        
        # Check if the user wants to shift empty cells
        option = input("Do you want to shift empty cells? (yes/no) ")
        if option == "yes":
            excel_workbook = openpyxl.load_workbook(output_excel_path)
            for worksheet_name in excel_workbook.sheetnames:
                excel_worksheet = excel_workbook[worksheet_name]
                for row in excel_worksheet.iter_rows():
                    empty_cell_list = []
                    for current_cell in row:
                        if current_cell.value is None:
                            empty_cell_list.append(current_cell)
                        else:
                            for empty_cell_to_fill in empty_cell_list:
                                empty_cell_to_fill.value = current_cell.value
                                current_cell.value = None
                                empty_cell_list.remove(empty_cell_to_fill)
                                empty_cell_list.append(current_cell)
                                break
            excel_workbook.save(output_excel_path)

        print(f"PDF converted to Excel: {output_excel_path}")

def split_pdf(filename, page_ranges, output_filename):
    try:
        pdf = PyPDF2.PdfReader(filename)
        output_pdf = PyPDF2.PdfWriter()

        for range_string in page_ranges:
            pages = []
            ranges = range_string.split(",")
            for rng in ranges:
                if "-" in rng:
                    start, end = map(int, rng.split("-"))
                    pages.extend(range(start, end + 1))
                else:
                    pages.append(int(rng))
            for page_num in pages:
                output_pdf.add_page(pdf.pages[page_num - 1])

        with open(output_filename, "wb") as output_file:
            output_pdf.write(output_file)

    except Exception as e:
        print(f"An error occurred: {str(e)}")

def merge_pdf_files(output_filename, all_files):
    try:
        if all_files:
            files = sorted(glob.glob(os.path.join(os.getcwd(), '*.pdf')))
            if not files:
                print("Error: No PDF files found in the current directory.")
                return

            merger = PyPDF2.PdfMerger()
            for filename in files:
                merger.append(filename)
            merger.write(output_filename)
            merger.close()
        else:
            filenames = []
            while True:
                try:
                    filename = input("Enter the filename (or 'done' to finish): ")
                    if filename == "done":
                        break
                    if not os.path.exists(filename):
                        print(f"Error: File '{filename}' does not exist.")
                        continue
                    filenames.append(filename)
                except Exception as e:
                    print(f"An error occurred: {e}")

            if not filenames:
                print("No files specified.")
                return

            merger = PyPDF2.PdfMerger()
            for filename in filenames:
                merger.append(filename)
            merger.write(output_filename)
            merger.close()

    except Exception as e:
        print(f"An error occurred: {e}")

def pdf_to_word(pdf_paths):
    try:
        docx_paths = []
        for pdf_path in pdf_paths:
            try:
                docx_path = pdf_path.replace(".pdf", ".docx")
                pdf2docx.parse(pdf_path, docx_path)
                docx_paths.append(docx_path)
            except Exception as e:
                print(f"Error converting PDF to Word: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

def pdf_to_images(pdf_paths):
    try:
        for pdf_path in pdf_paths:
            try:
                #ignore warnings
                warnings.simplefilter('ignore', PIL.Image.DecompressionBombWarning)
                images = pdf2image.convert_from_path(pdf_path, dpi=1000)
                for idx, img in enumerate(images):
                    img.save(f'page_{idx + 1}.jpg', 'JPEG', quality=80)
            except Exception as e:
                print(f"Error converting PDF to images: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")